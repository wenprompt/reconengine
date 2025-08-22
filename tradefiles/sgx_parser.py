
import openpyxl
import csv
from datetime import datetime
from rich.console import Console

console = Console()

def parse_io_xlsx(input_path, output_path):
    """
    Parses an Excel file (io.xlsx), transforms the data, and writes it to a CSV file.
    Processes all sheets in the workbook.

    Args:
        input_path (str): Path to the input Excel file.
        output_path (str): Path to the output CSV file.
    """
    try:
        workbook = openpyxl.load_workbook(input_path)
    except FileNotFoundError:
        console.print(f"[bold red]Error:[/bold red] Input file not found at {input_path}")
        return

    output_rows = []
    output_header = [
        "traderid", "tradedate", "tradetime", "productid", "productname", 
        "productgroupid", "exchangegroupid", "brokergroupid", "exchclearingacctid", 
        "quantitylots", "quantityunits", "unit", "price", "contractmonth", 
        "strike", "specialComms", "spread", "b/s", "put/call", "RMKS", "BKR"
    ]
    output_rows.append(output_header)

    # Process all sheets in the workbook
    total_trades = 0
    for sheet_name in workbook.sheetnames:
        console.print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Skip empty sheets or sheets with insufficient rows
        if sheet.max_row < 3:
            console.print(f"Skipping sheet '{sheet_name}' - insufficient data")
            continue

        header = [cell.value for cell in sheet[1]]
        
        # Find the contract months and their column indices
        contract_months = {}
        for i, cell_value in enumerate(header):
            if cell_value and isinstance(cell_value, str) and cell_value != "Counterparty":
                contract_months[cell_value] = i

        if not contract_months:
            console.print(f"Skipping sheet '{sheet_name}' - no contract month columns found")
            continue

        sheet_trades = 0
        for row_index in range(3, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_index]]
            if not any(row_values):
                continue

            for contract_month, start_col_idx in contract_months.items():
                quantity = row_values[start_col_idx]
                price = row_values[start_col_idx + 1]
                counterparty = row_values[start_col_idx + 2]
                time_val = row_values[start_col_idx + 3]

                if not all([quantity, price, counterparty, time_val]):
                    continue

                # Transformations
                try:
                    quantity_float = float(quantity)
                    quantity_lots = abs(quantity_float)
                    quantity_units = quantity_lots * 100
                    bs = 'B' if quantity_float > 0 else 'S'
                except (ValueError, TypeError):
                    continue

                spread = 'S' if "spread" in str(counterparty).lower() else ''
                counterparty_first_word = str(counterparty).split(" ")[0]

                tradetime = ""
                if hasattr(time_val, 'hour'):
                    tradetime = time_val.strftime('%H:%M:%S')
                elif isinstance(time_val, datetime.datetime):
                    tradetime = time_val.strftime('%H:%M:%S')
                elif isinstance(time_val, str):
                    try:
                        # Try parsing common time formats
                        if "am" in time_val.lower() or "pm" in time_val.lower():
                            dt_obj = datetime.strptime(time_val, '%I:%M:%S %p')
                        else:
                            dt_obj = datetime.strptime(time_val, '%H:%M:%S')
                        tradetime = dt_obj.strftime('%H:%M:%S')
                    except ValueError:
                        pass # Keep tradetime empty if parsing fails

                output_row = {
                    "traderid": "",
                    "tradedate": "",
                    "tradetime": tradetime,
                    "productid": "",
                    "productname": "FE",
                    "productgroupid": "",
                    "exchangegroupid": 1,
                    "brokergroupid": 3,
                    "exchclearingacctid": 2,
                    "quantitylots": quantity_lots,
                    "quantityunits": quantity_units,
                    "unit": "",
                    "price": price,
                    "contractmonth": contract_month,
                    "strike": "",
                    "specialComms": "",
                    "spread": spread,
                    "b/s": bs,
                    "put/call": "",
                    "RMKS": counterparty_first_word,
                    "BKR": ""
                }
                output_rows.append([output_row[h] for h in output_header])
                sheet_trades += 1

        console.print(f"Extracted {sheet_trades} trades from sheet '{sheet_name}'")
        total_trades += sheet_trades

    console.print(f"[bold blue]Total trades extracted from all sheets: {total_trades}[/bold blue]")

    with open(output_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(output_rows)

if __name__ == '__main__':
    input_file = 'tradefiles/input_traders/io.xlsx'
    output_file = 'tradefiles/output_traders/sourceTraders.csv'
    parse_io_xlsx(input_file, output_file)
    console.print(f"[bold green]Successfully parsed {input_file} and created {output_file}[/bold green]")
